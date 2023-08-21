import openpyxl
from openpyxl.styles import NamedStyle

source_filename = "C:/Users/ACER/Documents/Citogenetika/Python/Source.xlsx"
destination_filename = "C:/Users/ACER/Documents/Citogenetika/Python/Target.xlsx"

def copy_data(source_filename, destination_filename):
    # Load the source and destination workbooks
    source_workbook = openpyxl.load_workbook(source_filename)
    destination_workbook = openpyxl.load_workbook(destination_filename)
    
    # Select the desired sheets in both workbooks
    source_sheet = source_workbook.active
    destination_sheet = destination_workbook.active

    # Copy data from source to destination
    for row in source_sheet.iter_rows(values_only=True):
        destination_sheet.append(row)
    
    # Save the changes to the destination workbook
    destination_workbook.save(destination_filename)
  
# Specify the file names
source_file = 'Source.xlsx'
destination_file = 'Target.xlsx'

# Call the function to copy data
copy_data(source_file, destination_file)

print("Data copied successfully!")