import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

source_filename = "C:/Users/ACER/Documents/Citogenetika/Python/Source.xlsx"
filename = "C:/Users/ACER/Documents/Citogenetika/Python/Target.xlsx"

# Determine how many rows need formatting after copying data
source_workbook = openpyxl.load_workbook(source_filename)
source_sheet = source_workbook.active
row_count = source_sheet.max_row
dest_workbook = openpyxl.load_workbook(filename)
dest_sheet = dest_workbook.active
dest_row_count = dest_sheet.max_row
start_row = dest_row_count - row_count + 1

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
            
def format_first_cell_of_rows(dest_sheet, row_count):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    
    # Select the desired sheet
    sheet = workbook.active

    # Apply the style to the first cell of each of the first N rows
    for row_idx in range(sheet.max_row, sheet.max_row - row_count, -1):
        cell = sheet.cell(row=row_idx, column=1)
        cell.number_format = 'dd/mm/yy'
        cell.alignment = Alignment(horizontal = 'left')
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Save the changes to the workbook
    workbook.save('Target.xlsx')

# Specify the file name, sheet name, and the number of rows to format
excel_file = 'Target.xlsx'


# Call the function to format the first cells of the specified rows
format_first_cell_of_rows(excel_file, row_count)

print(f"Formatting applied to the first cells of the last {row_count} rows.")
