import openpyxl
import pandas as pd
import sys

if len(sys.argv) > 2:
    source_file = sys.argv[1]
    destination_file = sys.argv[2]

    def copy_columns_with_mapping(source_file, destination_file, column_mapping):
        # Load the source workbook and select the desired sheet
        source_workbook = openpyxl.load_workbook(source_file)
        source_sheet = source_workbook.active
          
        # Load the destination workbook and select the desired sheet
        destination_workbook = openpyxl.load_workbook(destination_file)
        destination_sheet = destination_workbook.active

        # Initialize a dictionary to store the current row for each destination column
        current_rows = {destination_header: destination_sheet.max_row + 1 for destination_header in column_mapping.values()}

        # Iterate through the source columns and copy data to destination columns
        for source_header, destination_header in column_mapping.items():
            source_column = None
            destination_column = None

            # Find source column by header name
            for cell in source_sheet[1]:
                if cell.value == source_header:
                    source_column = cell.column

            # Find destination column by header name
            for cell in destination_sheet[1]:
                if cell.value == destination_header:
                    destination_column = cell.column

            if source_column and destination_column:
                for row_idx in range(2, source_sheet.max_row + 1):
                    source_cell = source_sheet.cell(row=row_idx, column=source_column)
                    destination_cell = destination_sheet.cell(row=current_rows[destination_header], column=destination_column)
                    destination_cell.value = source_cell.value
                    current_rows[destination_header] += 1

        # Save the changes to the destination workbook
        destination_workbook.save(destination_file)

    # Define column mapping (source header: destination header)
    column_mapping = {
        'GENE': 'gene',
        'AFDP': 'depth',
        'c.DNA': 'c.DNA',
        'protein': 'protein',
        'Variant Allele Freq': 'var_percent',
        'Ref': 'ref',
        'Alt': 'alt',
        'NM': 'refSeqId',
        'Date': 'Datum',
        'Mintaszam': 'MG/LIMS_Mintaszam'
    
        # Add more mappings as needed
    }

    # Call the function to copy columns with mapping
    copy_columns_with_mapping(source_file, destination_file, column_mapping)

    print("data_copy_by_headers ran successfully")