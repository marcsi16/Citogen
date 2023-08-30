import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
import sys

def format_first_cell_of_rows(dest_sheet, row_count):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    # Apply the style to the first cell of each of the last N rows
    for row_idx in range(dest_sheet.max_row, dest_sheet.max_row - row_count, -1):
        cell = dest_sheet.cell(row=row_idx, column=1)
        cell.number_format = 'dd/mm/yy'
        cell.alignment = Alignment(horizontal='right')
        for col_idx in range(1, dest_sheet.max_column + 1):
            cell = dest_sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

if len(sys.argv) == 3:  # You need to provide both source and destination file paths
    source_file = sys.argv[1]
    destination_file = sys.argv[2]
    
    # Count rows in source excel
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook.active
    row_count = source_sheet.max_row

    dest_workbook = openpyxl.load_workbook(destination_file)
    dest_sheet = dest_workbook.active
    dest_row_count = dest_sheet.max_row
    start_row = dest_row_count - row_count + 1

    # Call the function 
    format_first_cell_of_rows(dest_sheet, row_count)

    # Save the changes to the destination workbook
    dest_workbook.save(destination_file)

    print(f"Formatting applied to the first cells of the last {row_count} rows.")
