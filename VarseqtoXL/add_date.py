from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys


if len(sys.argv) == 2:
    source_file = sys.argv[1]
    
    # Load the existing workbook
    workbook = load_workbook(filename=source_file)
    sheet = workbook.active

    # Get today's date
    today_date = datetime.now().date()

    # Determine the next available column letter
    next_column_letter = get_column_letter(sheet.max_column + 1)

    # Add today's date to the header row
    new_column_header = "Date"
    sheet[f"{next_column_letter}1"] = new_column_header

    # Add today's date to each row in the new column
    for row_idx in range(2, sheet.max_row + 1):
        sheet[f"{next_column_letter}{row_idx}"] = today_date

    # Save the modified workbook
    workbook.save(source_file)
 
    print("add_date.py ran succesfully")