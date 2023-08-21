import openpyxl
from openpyxl.styles import NamedStyle

source_filename = "C:/Users/ACER/Documents/Citogenetika/Python/Source.xlsx"
filename = "C:/Users/ACER/Documents/Citogenetika/Python/Target.xlsx"


# Determine how many rows need formatting after copying data
source_workbook = openpyxl.load_workbook(source_filename)
source_sheet = source_workbook.active
row_count = source_sheet.max_row
dest_workbook = openpyxl.load_workbook(filename)
dest_sheet = dest_workbook.active
dest_row_count = dest_sheet.max_row
print ("Number of rows in source worksheet: ", row_count)
print ("Number of rows in destination worksheet: ", dest_row_count)
print (dest_workbook.named_styles)